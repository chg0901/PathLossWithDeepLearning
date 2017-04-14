# -*- coding: utf-8 -*-
# __author__ = "CHG"
# Email: kwchenghong@gmail.com
'''
将数据分层，使可视化更加好看

'''

from openpyxl import Workbook  # write

from openpyxl import load_workbook  # read

# wbr = load_workbook("AP2delete.xlsx")
wbr = load_workbook("AP1_change_style_delete.xlsx")

wbw = Workbook()

sheetnames = wbr.get_sheet_names()
print(sheetnames)
r = len(sheetnames)

newSheet = wbw.create_sheet('(x,y,rssi)')
newSheet.cell(row=1, column=1, value='x')
newSheet.cell(row=1, column=2, value='y')
newSheet.cell(row=1, column=3, value='value')
ij = 2
for rnum in range(0, r):  # sheets loop r
    print(rnum)
    sheet = wbr.get_sheet_by_name(sheetnames[rnum])
    print(sheet.title)
    maxR = sheet.max_row
    maxC = sheet.max_column
    print('maxC', maxC)
    print('maxR', maxR)

    for i in range(2, maxR + 1):  # row loop maxR + 1

        xValue = sheet.cell(row=i, column=1).value
        yValue = sheet.cell(row=i, column=2).value
        print('xValue', xValue)
        print('yValue', yValue)

        for j in range(3, maxC + 1):  # column loop
            oldValue = sheet.cell(row=i, column=j).value
            if oldValue is None:
                wbw.save('AP1_coordinate,rssi.xlsx')
                break
            newSheet.cell(row=ij, column=1, value=xValue)
            newSheet.cell(row=ij, column=2, value=yValue)
            newSheet.cell(row=ij, column=3, value=oldValue)
            ij += 1
            wbw.save('AP1_coordinate,rssi.xlsx')


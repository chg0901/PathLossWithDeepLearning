# -*- coding: utf-8 -*-
# __author__ = "CHG"
# Email: kwchenghong@gmail.com

import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D   # for projection = 3d
from openpyxl import load_workbook  # read

wbr = load_workbook("AP2_coordinate,rssi.xlsx")

sheetnames = wbr.get_sheet_names()
print(sheetnames)
r = len(sheetnames)
sheet1 = wbr.get_sheet_by_name(sheetnames[0])

fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')

print(sheet1)

# # For each set of style and range settings, plot n random points in the box
# # defined by x in [23, 32], y in [0, 100], z in [zlow, zhigh].
# for c, m, zlow, zhigh in [('r', 'o', -70, -50), ('y', '*', -49, -35), ('b', '^', -34, -20)]:
#     xs = randrange(n, 23, 32)
#     ys = randrange(n, 0, 100)
#     zs = randrange(n, zlow, zhigh)
#     ax.scatter(xs, ys, zs, c=c, marker=m)
#
# ax.set_xlabel('X Label')
# ax.set_ylabel('Y Label')
# ax.set_zlabel('Z Label')
#
# plt.show()

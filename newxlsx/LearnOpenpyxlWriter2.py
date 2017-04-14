# -*- coding: utf-8 -*-
# __author__ = "CHG"
# Email: kwchenghong@gmail.com

from openpyxl.workbook import Workbook

from openpyxl.writer.excel import ExcelWriter

wb = Workbook()

ew = ExcelWriter(workbook=wb)
dest_filename = r'empty_book.xlsx'

sheet1 = wb.worksheets[0]
sheet1.title = 'range names'
sheet1.cell('C1').value = u'aizhaojie'
sheet1.cell(2, 'A', 10)

sheet2 = wb.create_sheet('sheet2', 1)
for i in range(1, 101):
    for j in range(1, 101):
        sheet2.cell(row=i, column=j)

ew.save(filename=dest_filename)

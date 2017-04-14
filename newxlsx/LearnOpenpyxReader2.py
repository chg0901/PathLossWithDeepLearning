# -*- coding: utf-8 -*-
# __author__ = "CHG"
# Email: kwchenghong@gmail.com

from openpyxl.reader.excel import load_workbook

wb = load_workbook('excel2.xlsx')

sheetnames = wb.get_sheet_names()
print(sheetnames)

print('sheetname2:')
print(wb.sheetnames)

# for loop
for sheet in wb:
    print(sheet.title)

# 取得第一张表
sheet1 = wb.get_sheet_by_name(sheetnames[0])

# 获取特定的worksheet
ws = wb.get_sheet_by_name('(X,1)')

# 获取表名
print(sheet1.title)

# 获取表的行数，表的列数

# # 旧的方法
# rows = sheet1.get_highest_row()
# columns = sheet1.get_highest_column()
# print 'rows and column'
# print rows
# print columns

rows2 = sheet1.max_row
columns2 = sheet1.max_column
print('rows and column')
print(rows2)
print(columns2)

# 单元格读取
print(sheet1.cell('F10').value)
# 只能读取数值数据？
# print sheet1.cell(0, 0).value
# AttributeError: 'int' object has no attribute 'upper'


# # 写文件只能通过坐标操作
# 如要向单元格C1写数据，就要用类似sheet1.cell('C1').value = something
# 的方式

print('(10,10)')
print(sheet1.title)

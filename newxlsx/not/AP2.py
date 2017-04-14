# -*- coding: utf-8 -*-
# __author__ = "CHG"
# Email: kwchenghong@gmail.com

from openpyxl import Workbook  # write
from openpyxl import load_workbook  # read

wbr = load_workbook("AP2delete.xlsx")
print(wbr.get_sheet_names())
print()

wbw = Workbook()

sheetnames = wbr.get_sheet_names()

newSheet = wbw.create_sheet('(x,y,rssi)')
r = len(wbr.sheetnames)
for rnum in range(0, r - 1):
    sheet = wbr.get_sheet_by_name(sheetnames[rnum]):
    print(type(sheet))  # <class 'tuple'> 'tuple' object has no attribute 'max_row'

    print(type(wbr.get_sheet_by_name(sheetnames[r])))
    print(sheet.title)

    for i in range(2, sheet.max_row):
        for j in range(3, sheet.max_column):
            # if  # 判断单元格不为空
            for ij in i * j:
                newSheet.cell(row=ij, column=1, value=sheet.cell(i, j).value)

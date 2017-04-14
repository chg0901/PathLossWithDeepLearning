# -*- coding: utf-8 -*-
# __author__ = "CHG"
# Email: kwchenghong@gmail.com

from openpyxl import Workbook  # write
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook  # read

wbr = load_workbook("AP2delete2.xlsx")
wbw = Workbook()

sheetnames = wbr.get_sheet_names()
print(sheetnames)
r = len(sheetnames)

ewb = ExcelWriter(workbook=wbw)  # 新建一个ExcelWriter，用来写wb1
dest_filename = r'(x,y,rssi).xlsx'

newSheet = wbw.worksheets[0]
newSheet.title = u'(x,y,rssi)'

newSheet.cell(row=1, column=1, value='x')
newSheet.cell(row=1, column=2, value='y')
newSheet.cell(row=1, column=3, value='value')

sheet = wbr.get_sheet_by_name(sheetnames[0])

for rnum in range(0, r):  # sheets loop
    sheet = wbr.get_sheet_by_name(sheetnames[rnum])
    maxR = sheet.max_row
    maxC = sheet.max_column
    print(maxC)
    print(maxR)
    for ij in range(2, maxR * maxC + 1):  # cell loop
        for i in range(2, maxR + 1):  # row loop
            xValue = sheet.cell(row=i, column=1).value
            yValue = sheet.cell(row=i, column=2).value
            print(xValue)
            print(yValue)

            for j in range(3, maxC + 1):  # column loop
                newSheet.cell(row=ij, column=1).value = xValue
                newSheet.cell(row=ij, column=2).value = yValue
                newSheet.cell(row=ij, column=3).value = sheet.cell(row=i, column=j).value
                print(type(sheet.cell(row=i, column=j).value))
ewb.save(filename=dest_filename)

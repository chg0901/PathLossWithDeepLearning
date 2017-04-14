# -*- coding: utf-8 -*-
# __author__ = "CHG"
# Email: kwchenghong@gmail.com

from openpyxl import Workbook  # write
from openpyxl import load_workbook  # read

wb = Workbook()

ws = wb.active  # create one Sheet

ws1 = wb.create_sheet("Mysheet1")
ws2 = wb.create_sheet("Mysheet0", 0)
ws.title = '123'

print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

ws['A4'] = 4

d = ws.cell(row=4, column=4, value=10)

for i in range(1, 101):
    for j in range(1, 101):
        ws.cell(row=i, column=j, value=i * j)

wb.save('mytable.xlsx')

wb2 = load_workbook("mytable.xlsx")
print(wb2.get_sheet_names())
print(wb2.sheetnames)

